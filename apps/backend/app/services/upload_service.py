"""
Upload service — handles file upload, validation, parsing and storage.

Fixes applied
-------------
1. _detect_file_type: when xlrd fails on a .xls file (because Tally saves
   xlsx-format files with a .xls extension), falls back to openpyxl.
2. parse_gstr2a is called with the actual file bytes; the parser itself
   already handles the openpyxl-first / xlrd-fallback routing correctly.
3. Gstr2ARecord creation saves `invoice_number` from the parsed invoice.
"""

import logging
import os
import time
from datetime import datetime, timezone
from io import BytesIO
from typing import Iterable

import openpyxl
import xlrd

from app.config.settings import settings
from app.models.gstr2a import Gstr2ARecord
from app.models.gstr2b import Gstr2BRecord
from app.models.user import User, Gstr2AFileRef, Gstr2BFileRef
from app.parsers.gstr2a_parser import parse_gstr2a
from app.parsers.gstr2b_parser import parse_gstr2b
from app.schemas.api import UploadResponse
from app.services import s3_service
from app.utils.date_helpers import parse_gst_date, to_period

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}

MONTH_NAME_MAP = {
    "january": 1,  "february": 2,  "march": 3,    "april": 4,
    "may": 5,       "june": 6,      "july": 7,     "august": 8,
    "september": 9, "october": 10,  "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _gstr2b_period_to_yyyy_mm(tax_period: str, financial_year: str) -> str:
    if not tax_period or not financial_year:
        return ""
    month_num = MONTH_NAME_MAP.get(tax_period.strip().lower())
    if not month_num:
        return ""
    try:
        start_year = int(financial_year.strip().split("-")[0])
    except (ValueError, IndexError):
        return ""
    year = start_year if month_num >= 4 else start_year + 1
    return f"{year}-{month_num:02d}"


def _safe_period_from_date(date_str: str) -> str:
    if not date_str:
        return ""
    try:
        return to_period(parse_gst_date(date_str))
    except (ValueError, TypeError):
        return ""


def _collect_non_empty(values: Iterable[str]) -> set[str]:
    return {v for v in values if v}


# ── File type detection ───────────────────────────────────────────────────────

def _sniff_with_openpyxl(file_bytes: bytes) -> str | None:
    """
    Try to detect GSTR-2A vs GSTR-2B using openpyxl.
    Returns "GSTR_2A", "GSTR_2B", or None if detection is inconclusive.
    """
    try:
        wb = openpyxl.load_workbook(
            BytesIO(file_bytes), read_only=True, data_only=True
        )
        sheet_names_lower = [s.lower().strip() for s in wb.sheetnames]

        # GSTR-2B always has a "Read me" sheet
        if "read me" in sheet_names_lower:
            wb.close()
            return "GSTR_2B"

        # Scan first 10 rows of first sheet for keywords
        ws = wb.worksheets[0]
        rows_checked = 0
        result = None
        for row in ws.iter_rows(values_only=True):
            if rows_checked >= 10:
                break
            for cell in row:
                if cell and isinstance(cell, str):
                    val = cell.upper()
                    if "GSTR-2B" in val:
                        result = "GSTR_2B"
                        break
                    if "GSTR-2A" in val:
                        result = "GSTR_2A"
                        break
            if result:
                break
            rows_checked += 1

        wb.close()
        return result
    except Exception as e:
        logger.debug(f"[upload] openpyxl sniff failed: {e}")
        return None


def _sniff_with_xlrd(file_bytes: bytes) -> str | None:
    """
    Try to detect GSTR-2A vs GSTR-2B using xlrd (for genuine .xls files).
    Returns "GSTR_2A", "GSTR_2B", or None if detection fails or is inconclusive.
    """
    try:
        wb           = xlrd.open_workbook(file_contents=file_bytes)
        sheet_names  = [s.lower().strip() for s in wb.sheet_names()]

        if "read me" in sheet_names:
            return "GSTR_2B"

        ws = wb.sheet_by_index(0)
        for row_idx in range(min(10, ws.nrows)):
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_TEXT:
                    val = cell.value.upper()
                    if "GSTR-2B" in val:
                        return "GSTR_2B"
                    if "GSTR-2A" in val:
                        return "GSTR_2A"
        return None
    except Exception as e:
        logger.debug(f"[upload] xlrd sniff failed: {e}")
        return None


def _detect_file_type(file_bytes: bytes, file_name: str) -> str:
    """
    Detect whether the file is GSTR-2A or GSTR-2B.

    Tally frequently saves xlsx-format files with a .xls extension.
    xlrd v2 cannot read xlsx-format files at all. So the strategy is:

      .xlsx / .xlsm → openpyxl only
      .xls          → try xlrd first; if it fails, fall back to openpyxl
      unknown ext   → try openpyxl first, then xlrd

    Falls back to "GSTR_2A" when all detection attempts are inconclusive.
    """
    ext = os.path.splitext(file_name)[1].lower()

    if ext in {".xlsx", ".xlsm"}:
        return _sniff_with_openpyxl(file_bytes) or "GSTR_2A"

    if ext == ".xls":
        # Try xlrd first (genuine .xls binary format)
        result = _sniff_with_xlrd(file_bytes)
        if result:
            return result
        # xlrd failed — file is probably xlsx-format with .xls extension (Tally)
        logger.info(
            f"[upload] xlrd could not read '{file_name}' "
            "(likely xlsx-format with .xls extension). Falling back to openpyxl."
        )
        return _sniff_with_openpyxl(file_bytes) or "GSTR_2A"

    # Unknown extension: try both
    return _sniff_with_openpyxl(file_bytes) or _sniff_with_xlrd(file_bytes) or "GSTR_2A"


# ── Idempotent record deletion ────────────────────────────────────────────────

async def _delete_existing_gstr2a_records(
    user_id: str,
    period_start: str,
    period_end: str,
    invoice_dates: list[str],
) -> None:
    if period_start or period_end:
        await Gstr2ARecord.find(
            Gstr2ARecord.user_id      == user_id,
            Gstr2ARecord.period_start == period_start,
            Gstr2ARecord.period_end   == period_end,
        ).delete()
        return
    target_periods = _collect_non_empty(
        _safe_period_from_date(d) for d in invoice_dates
    )
    if not target_periods:
        return
    existing = await Gstr2ARecord.find(Gstr2ARecord.user_id == user_id).to_list()
    for record in existing:
        if _safe_period_from_date(record.date) in target_periods:
            await record.delete()


async def _delete_existing_gstr2b_records(
    user_id: str,
    financial_year: str,
    tax_period: str,
    invoice_dates: list[str],
) -> None:
    if financial_year or tax_period:
        await Gstr2BRecord.find(
            Gstr2BRecord.user_id        == user_id,
            Gstr2BRecord.financial_year == financial_year,
            Gstr2BRecord.tax_period     == tax_period,
        ).delete()
        return
    target_periods = _collect_non_empty(
        _safe_period_from_date(d) for d in invoice_dates
    )
    if not target_periods:
        return
    existing = await Gstr2BRecord.find(Gstr2BRecord.user_id == user_id).to_list()
    for record in existing:
        period = (
            _gstr2b_period_to_yyyy_mm(record.tax_period, record.financial_year)
            or _safe_period_from_date(record.invoice_date)
        )
        if period in target_periods:
            await record.delete()


# ── User file-reference management ───────────────────────────────────────────

async def _replace_gstr2a_file_ref(user_id: str, file_ref: Gstr2AFileRef) -> None:
    user = await User.find_one(User.user_id == user_id)
    if user is None:
        return
    user.gstr2a_files = [
        ref for ref in user.gstr2a_files
        if ref.period != file_ref.period and ref.file_name != file_ref.file_name
    ]
    user.gstr2a_files.append(file_ref)
    user.updated_at = datetime.now(timezone.utc)
    await user.save()


async def _replace_gstr2b_file_ref(user_id: str, file_ref: Gstr2BFileRef) -> None:
    user = await User.find_one(User.user_id == user_id)
    if user is None:
        return
    user.gstr2b_files = [
        ref for ref in user.gstr2b_files
        if not (
            ref.file_name == file_ref.file_name
            or (
                ref.tax_period     == file_ref.tax_period
                and ref.financial_year == file_ref.financial_year
            )
        )
    ]
    user.gstr2b_files.append(file_ref)
    user.updated_at = datetime.now(timezone.utc)
    await user.save()


# ── Main handler ──────────────────────────────────────────────────────────────

async def handle_upload(file_bytes: bytes, file_name: str, user_id: str) -> UploadResponse:
    """Validate, detect type, parse, and store an uploaded Excel file."""

    ext = os.path.splitext(file_name)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"Invalid file type '{ext}'. Only .xlsx and .xls are allowed."
        )

    max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if len(file_bytes) > max_bytes:
        raise ValueError(
            f"File size exceeds maximum of {settings.MAX_UPLOAD_SIZE_MB} MB."
        )

    file_type = _detect_file_type(file_bytes, file_name)
    file_id   = f"{file_type.lower()}_{user_id}_{int(time.time())}"
    logger.info(f"[upload] Detected file_type={file_type} for '{file_name}'")

    s3_key = f"gstr_uploads/{user_id}/{file_id}{ext}"
    s3_service.upload_file(file_bytes, s3_key)

    # ── GSTR-2A ──────────────────────────────────────────────────────────────
    if file_type == "GSTR_2A":
        result = parse_gstr2a(file_bytes, file_name)

        await _delete_existing_gstr2a_records(
            user_id       = user_id,
            period_start  = result.metadata.period_start,
            period_end    = result.metadata.period_end,
            invoice_dates = [inv.date for inv in result.invoices],
        )

        records = [
            Gstr2ARecord(
                user_id        = user_id,
                file_name      = file_name,
                period_start   = result.metadata.period_start,
                period_end     = result.metadata.period_end,
                company_name   = result.metadata.company_name,
                gstin          = result.metadata.gstin,
                date           = inv.date,
                particulars    = inv.particulars,
                party_gstin    = inv.party_gstin,
                vch_type       = inv.vch_type,
                vch_no         = inv.vch_no,
                invoice_number = inv.invoice_number,   # ← official supplier invoice
                taxable_amount = round(inv.taxable_amount, 2),
                igst           = round(inv.igst, 2),
                cgst           = round(inv.cgst, 2),
                sgst_utgst     = round(inv.sgst_utgst, 2),
                cess           = round(inv.cess, 2),
                tax_amount     = round(inv.tax_amount, 2),
                invoice_amount = round(inv.invoice_amount, 2),
            )
            for inv in result.invoices
        ]
        if records:
            await Gstr2ARecord.insert_many(records)

        file_ref = Gstr2AFileRef(
            file_id     = file_id,
            file_name   = file_name,
            period      = f"{result.metadata.period_start} to {result.metadata.period_end}",
            uploaded_at = datetime.now(timezone.utc),
        )
        await _replace_gstr2a_file_ref(user_id, file_ref)

        return UploadResponse(
            success        = True,
            message        = f"GSTR-2A uploaded — {len(records)} records parsed.",
            file_id        = file_id,
            file_name      = file_name,
            file_type      = "GSTR_2A",
            records_parsed = len(records),
        )

    # ── GSTR-2B ──────────────────────────────────────────────────────────────
    else:
        result = parse_gstr2b(file_bytes, file_name)

        await _delete_existing_gstr2b_records(
            user_id        = user_id,
            financial_year = result.metadata.financial_year,
            tax_period     = result.metadata.tax_period,
            invoice_dates  = [inv.invoice_date for inv in result.b2b_invoices],
        )

        records = [
            Gstr2BRecord(
                user_id                        = user_id,
                file_name                      = file_name,
                financial_year                 = result.metadata.financial_year,
                tax_period                     = result.metadata.tax_period,
                buyer_gstin                    = result.metadata.buyer_gstin,
                legal_name                     = result.metadata.legal_name,
                trade_name                     = result.metadata.trade_name,
                date_of_generation             = result.metadata.date_of_generation,
                sheet_name                     = inv.sheet_name,
                supplier_gstin                 = inv.supplier_gstin,
                supplier_trade_name            = inv.supplier_trade_name,
                invoice_number                 = inv.invoice_number,
                invoice_type                   = inv.invoice_type,
                invoice_date                   = inv.invoice_date,
                invoice_value                  = round(inv.invoice_value, 2),
                place_of_supply                = inv.place_of_supply,
                supply_attracts_reverse_charge = inv.supply_attracts_reverse_charge,
                tax_rate                       = inv.tax_rate,
                taxable_value                  = round(inv.taxable_value, 2),
                integrated_tax                 = round(inv.integrated_tax, 2),
                central_tax                    = round(inv.central_tax, 2),
                state_ut_tax                   = round(inv.state_ut_tax, 2),
                cess                           = round(inv.cess, 2),
                gstr1_period                   = inv.gstr1_period,
                gstr1_filing_date              = inv.gstr1_filing_date,
                itc_availability               = inv.itc_availability,
                itc_unavailable_reason         = inv.itc_unavailable_reason,
                applicable_tax_rate_percent    = inv.applicable_tax_rate_percent,
                source                         = inv.source,
                irn                            = inv.irn,
                irn_date                       = inv.irn_date,
            )
            for inv in result.b2b_invoices
        ]
        if records:
            await Gstr2BRecord.insert_many(records)

        file_ref = Gstr2BFileRef(
            file_id        = file_id,
            file_name      = file_name,
            tax_period     = result.metadata.tax_period,
            financial_year = result.metadata.financial_year,
            uploaded_at    = datetime.now(timezone.utc),
        )
        await _replace_gstr2b_file_ref(user_id, file_ref)

        return UploadResponse(
            success        = True,
            message        = f"GSTR-2B uploaded — {len(records)} records parsed.",
            file_id        = file_id,
            file_name      = file_name,
            file_type      = "GSTR_2B",
            records_parsed = len(records),
        )
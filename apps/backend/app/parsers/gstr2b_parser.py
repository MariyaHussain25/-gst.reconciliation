"""
GSTR-2B Excel parser using openpyxl.
Rewrite of gstr2b.parser.ts with identical logic.
"""

import re
from io import BytesIO
from typing import Optional
import openpyxl
from pydantic import BaseModel


class Gstr2BMetadata(BaseModel):
    financial_year: str = ""
    tax_period: str = ""
    buyer_gstin: str = ""
    legal_name: str = ""
    trade_name: str = ""
    date_of_generation: str = ""


class Gstr2BInvoice(BaseModel):
    sheet_name: str = ""
    supplier_gstin: str = ""
    supplier_trade_name: str = ""
    invoice_number: str = ""
    invoice_type: str = ""
    invoice_date: str = ""
    invoice_value: float = 0.0
    place_of_supply: str = ""
    supply_attracts_reverse_charge: str = ""
    tax_rate: Optional[float] = None
    taxable_value: float = 0.0
    integrated_tax: float = 0.0
    central_tax: float = 0.0
    state_ut_tax: float = 0.0
    cess: float = 0.0
    gstr1_period: Optional[str] = None
    gstr1_filing_date: Optional[str] = None
    itc_availability: str = ""
    itc_unavailable_reason: Optional[str] = None
    applicable_tax_rate_percent: Optional[float] = None
    source: str = ""
    irn: Optional[str] = None
    irn_date: Optional[str] = None


class ItcSummary(BaseModel):
    part_a: dict = {}
    part_b: dict = {}


class Gstr2BParseResult(BaseModel):
    metadata: Gstr2BMetadata
    b2b_invoices: list[Gstr2BInvoice]
    itc_available_summary: ItcSummary
    itc_not_available_summary: ItcSummary


CORE_COLUMNS = ("party_gstin", "invoice_number", "invoice_date", "taxable_amount")

GST_HEADER_KEYWORDS = ("gstin", "tax", "invoice", "amount", "date", "particulars", "supplier")

COLUMN_ALIASES: dict[str, list[str]] = {
    "party_gstin": ["GSTIN of supplier", "Supplier GSTIN", "GSTIN/UIN"],
    "supplier_name": ["Trade/Legal name", "Supplier Trade Name", "Particulars"],
    "invoice_number": ["Invoice number", "Invoice No.", "Vch No."],
    "invoice_type": ["Invoice type", "Vch Type"],
    "invoice_date": ["Invoice date", "Invoice Date", "Date"],
    "invoice_value": ["Invoice Value(₹)", "Invoice Value", "Invoice Amount"],
    "place_of_supply": ["Place of supply"],
    "reverse_charge": ["Supply Attracts Reverse Charge", "Reverse Charge"],
    "tax_rate": ["Rate(%)", "Tax Rate"],
    "taxable_amount": ["Taxable value", "Taxable Value", "Taxable Amount", "Taxable Income"],
    "igst": ["Integrated Tax", "IGST"],
    "cgst": ["Central Tax", "CGST"],
    "sgst": ["State/UT tax", "State/UT Tax", "State Tax", "SGST", "SGST/UTGST"],
    "cess": ["Cess", "Cess Amount"],
    "gstr1_period": ["GSTR-1/IFF/GSTR-5 Period"],
    "gstr1_filing_date": ["GSTR-1/IFF/GSTR-5 Filing Date"],
    "itc_availability": ["ITC Availability"],
    "reason": ["Reason"],
    "applicable_tax_rate_percent": ["Applicable % of Tax Rate"],
    "source": ["Source"],
    "irn": ["IRN"],
    "irn_date": ["IRN date"],
}


def safe_number(val) -> float:
    if val is None or val == "":
        return 0.0
    try:
        if isinstance(val, str):
            cleaned = val.replace("₹", "").replace(",", "").strip()
            return float(cleaned) if cleaned else 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_optional_number(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def safe_optional_str(val) -> Optional[str]:
    if val is None or safe_str(val) == "":
        return None
    return str(val).strip()


def _normalize_header(value) -> str:
    text = safe_str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _detect_header_row(all_rows: list[tuple]) -> int:
    best_idx = -1
    best_score = -1
    for i, row in enumerate(all_rows[:15]):
        if not row:
            continue
        score = 0
        for cell in row:
            normalized = _normalize_header(cell)
            if not normalized:
                continue
            if any(keyword in normalized for keyword in GST_HEADER_KEYWORDS):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score > 0 else -1


def _build_column_map(header_row: tuple) -> dict[str, int]:
    alias_lookup: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_lookup[_normalize_header(alias)] = canonical

    resolved: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        canonical = alias_lookup.get(normalized)
        if canonical and canonical not in resolved:
            resolved[canonical] = idx
    return resolved


def _extract_readme_metadata(ws) -> Gstr2BMetadata:
    """Extract metadata from 'Read me' sheet."""
    metadata = Gstr2BMetadata()
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = safe_str(row[0]).lower()
        value = safe_str(row[1]) if len(row) > 1 and row[1] is not None else ""
        if "financial year" in key:
            metadata.financial_year = value
        elif "tax period" in key:
            metadata.tax_period = value
        elif "gstin" in key:
            metadata.buyer_gstin = value
        elif "legal name" in key:
            metadata.legal_name = value
        elif "trade name" in key:
            metadata.trade_name = value
        elif "date of generation" in key:
            metadata.date_of_generation = value
    return metadata


def _parse_b2b_sheet(ws) -> list[Gstr2BInvoice]:
    """Parse the B2B sheet into invoice records."""
    all_rows = list(ws.iter_rows(values_only=True))
    header_row_idx = _detect_header_row(all_rows)

    if header_row_idx == -1:
        return []
    col_map = _build_column_map(all_rows[header_row_idx])
    if any(column not in col_map for column in CORE_COLUMNS):
        return []

    invoices: list[Gstr2BInvoice] = []
    for row in all_rows[header_row_idx + 1:]:
        if not row or all(cell is None for cell in row):
            continue

        def get_cell(col_name: str):
            idx = col_map.get(col_name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        supplier_gstin = safe_str(get_cell("party_gstin")).upper().replace(" ", "")
        invoice_number = safe_str(get_cell("invoice_number"))
        invoice_date = safe_str(get_cell("invoice_date"))
        if not supplier_gstin or not invoice_number or not invoice_date:
            continue

        taxable_value = safe_number(get_cell("taxable_amount"))
        integrated_tax = safe_number(get_cell("igst"))
        central_tax = safe_number(get_cell("cgst"))
        state_ut_tax = safe_number(get_cell("sgst"))
        cess = safe_number(get_cell("cess"))
        invoice_value = safe_number(get_cell("invoice_value"))
        if invoice_value == 0.0:
            invoice_value = round(taxable_value + integrated_tax + central_tax + state_ut_tax + cess, 2)

        invoice = Gstr2BInvoice(
            sheet_name="B2B",
            supplier_gstin=supplier_gstin,
            supplier_trade_name=safe_str(get_cell("supplier_name")),
            invoice_number=invoice_number,
            invoice_type=safe_str(get_cell("invoice_type")),
            invoice_date=invoice_date,
            invoice_value=invoice_value,
            place_of_supply=safe_str(get_cell("place_of_supply")),
            supply_attracts_reverse_charge=safe_str(get_cell("reverse_charge")),
            tax_rate=safe_optional_number(get_cell("tax_rate")),
            taxable_value=taxable_value,
            integrated_tax=integrated_tax,
            central_tax=central_tax,
            state_ut_tax=state_ut_tax,
            cess=cess,
            gstr1_period=safe_optional_str(get_cell("gstr1_period")),
            gstr1_filing_date=safe_optional_str(get_cell("gstr1_filing_date")),
            itc_availability=safe_str(get_cell("itc_availability")),
            itc_unavailable_reason=safe_optional_str(get_cell("reason")),
            applicable_tax_rate_percent=safe_optional_number(get_cell("applicable_tax_rate_percent")),
            source=safe_str(get_cell("source")),
            irn=safe_optional_str(get_cell("irn")),
            irn_date=safe_optional_str(get_cell("irn_date")),
        )
        invoices.append(invoice)

    return invoices


def _parse_itc_summary_sheet(ws) -> ItcSummary:
    """Parse ITC summary sheet into ItcSummary."""
    part_a: dict = {}
    part_b: dict = {}
    current_part = None
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        cell = safe_str(row[0]).lower()
        if "part a" in cell:
            current_part = "a"
        elif "part b" in cell:
            current_part = "b"
        elif current_part and row[0] is not None:
            key = safe_str(row[0])
            value = safe_number(row[1]) if len(row) > 1 else 0.0
            if current_part == "a":
                part_a[key] = value
            else:
                part_b[key] = value
    return ItcSummary(part_a=part_a, part_b=part_b)


def parse_gstr2b(file_bytes: bytes, file_name: str) -> Gstr2BParseResult:
    """Parse GSTR-2B Excel file and return structured data."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        # Extract metadata from "Read me" sheet
        metadata = Gstr2BMetadata()
        for sheet_name in wb.sheetnames:
            if sheet_name.lower().strip() == "read me":
                metadata = _extract_readme_metadata(wb[sheet_name])
                break

        # Parse B2B invoices
        b2b_invoices: list[Gstr2BInvoice] = []
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().upper() == "B2B":
                b2b_invoices = _parse_b2b_sheet(wb[sheet_name])
                break
        if not b2b_invoices and wb.sheetnames:
            # Graceful fallback for ERP exports where sheet is not literally named "B2B".
            b2b_invoices = _parse_b2b_sheet(wb[wb.sheetnames[0]])

        # Parse ITC summary sheets
        itc_available_summary = ItcSummary()
        itc_not_available_summary = ItcSummary()
        for sheet_name in wb.sheetnames:
            name_lower = sheet_name.lower().strip()
            if "itc available" in name_lower and "not" not in name_lower:
                itc_available_summary = _parse_itc_summary_sheet(wb[sheet_name])
            elif "itc not available" in name_lower:
                itc_not_available_summary = _parse_itc_summary_sheet(wb[sheet_name])

        return Gstr2BParseResult(
            metadata=metadata,
            b2b_invoices=b2b_invoices,
            itc_available_summary=itc_available_summary,
            itc_not_available_summary=itc_not_available_summary,
        )
    finally:
        wb.close()

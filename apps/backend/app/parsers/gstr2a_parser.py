"""
GSTR-2A Excel parser.
Supports .xlsx (openpyxl) and legacy .xls (xlrd ≥ 2.0).

NOTE: xlrd v2+ dropped .xlsx support entirely.
      .xlsx → openpyxl   (always)
      .xls  → xlrd       (always)
      unknown extension  → try openpyxl first, fall back to xlrd

Two invoice number columns are tracked separately:
  vch_no         ← "Vch No."        column (internal Tally voucher, e.g. "2513")
  invoice_number ← "invoice_number" column (official supplier invoice, e.g. "02103")

When only one column is present, invoice_number falls back to vch_no value.
"""

import re
from datetime import datetime, timezone, date as date_type
from io import BytesIO
from typing import Optional

import openpyxl
import xlrd
from pydantic import BaseModel


# ── Pydantic schemas ──────────────────────────────────────────────────────────

class Gstr2AMetadata(BaseModel):
    company_name: str = ""
    period_start: str = ""
    period_end: str = ""
    gstin: str = ""


class Gstr2AInvoice(BaseModel):
    date: str = ""
    particulars: str = ""
    party_gstin: str = ""
    vch_type: str = ""
    vch_no: str = ""            # internal Tally voucher number
    invoice_number: str = ""    # official supplier invoice number
    taxable_amount: float = 0.0
    igst: float = 0.0
    cgst: float = 0.0
    sgst_utgst: float = 0.0
    cess: float = 0.0
    tax_amount: float = 0.0
    invoice_amount: float = 0.0


class Gstr2AParseResult(BaseModel):
    metadata: Gstr2AMetadata
    invoices: list[Gstr2AInvoice]


# ── Utility functions ─────────────────────────────────────────────────────────

def safe_number(val) -> float:
    if val is None or val == "":
        return 0.0
    try:
        if isinstance(val, str):
            cleaned = val.replace("₹", "").replace(",", "").strip()
            return float(cleaned) if cleaned else 0.0
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def safe_str(val) -> str:
    """Convert any value to a clean string. Handles openpyxl date objects."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%-d-%b-%y")
    if isinstance(val, date_type):
        return datetime(val.year, val.month, val.day).strftime("%-d-%b-%y")
    return str(val).strip()


# ── Column alias map ──────────────────────────────────────────────────────────

CORE_COLUMNS = ("party_gstin", "invoice_date", "taxable_amount")

GST_HEADER_KEYWORDS = (
    "gstin", "tax", "invoice", "amount", "date", "particulars", "vch"
)

COLUMN_ALIASES: dict[str, list[str]] = {
    "party_gstin": [
        "Party GSTIN/UIN", "GSTIN/UIN", "GSTIN of supplier",
        "party_gstin", "gstin_uin",
    ],
    "vch_type": [
        "Vch Type", "Vch Typ", "Voucher Type",
        "vch_type", "voucher_type",
    ],
    # Internal Tally voucher number — "Vch No." column
    "vch_number": [
        "Vch No.", "Vch No", "Voucher No.", "Voucher No",
        "vch_no", "vch_number", "voucher_no",
    ],
    # Official supplier invoice number — "invoice_number" / "Invoice No." column
    "invoice_number": [
        "invoice_number", "Invoice No.", "Invoice No",
        "Invoice Number", "Invoice number", "Inv No.", "Inv No",
    ],
    "invoice_date": [
        "Date", "Invoice Date", "Invoice date", "Inv Date",
        "invoice_date", "date",
    ],
    "particulars": [
        "Particulars", "Vendor Name", "Supplier Name",
        "particulars", "vendor_name", "supplier_name",
    ],
    "taxable_amount": [
        "Taxable Amount", "Taxable Value", "Taxable Income", "Taxable",
        "Taxable Value (₹)", "Taxable Value(₹)",
        "taxable_amount", "taxable_value",
    ],
    "igst": [
        "IGST", "Integrated Tax", "Integrated Tax Amount", "Integrated Tax(₹)",
        "igst", "integrated_tax",
    ],
    "cgst": [
        "CGST", "Central Tax", "Central Tax Amount", "Central Tax(₹)",
        "cgst", "central_tax",
    ],
    "sgst": [
        "SGST/UTGST", "SGST", "State Tax", "State/UT Tax",
        "State Tax Amount", "State/UT Tax(₹)",
        "sgst", "sgst_utgst", "state_tax",
    ],
    "cess": [
        "Cess", "Cess Amount", "Cess(₹)",
        "cess", "cess_amount",
    ],
    "tax_amount": [
        "Tax Amount", "Total Tax", "Total Tax Amount",
        "tax_amount", "total_tax",
    ],
    "invoice_amount": [
        "Invoice Amount", "Invoice Value", "Invoice Value(₹)", "Invoice Value (₹)",
        "invoice_amount", "invoice_value",
    ],
}


def _normalize_header(value) -> str:
    text = safe_str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _detect_header_row(all_rows: list[tuple]) -> int:
    best_idx, best_score = -1, -1
    for i, row in enumerate(all_rows[:20]):
        if not row:
            continue
        score = sum(
            1 for cell in row
            if _normalize_header(cell)
            and any(kw in _normalize_header(cell) for kw in GST_HEADER_KEYWORDS)
        )
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx if best_score > 0 else -1


def _build_column_map(header_row: tuple) -> dict[str, int]:
    # Build alias → canonical lookup
    alias_lookup: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_lookup[_normalize_header(alias)] = canonical

    resolved: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        if not normalized:
            continue
        # Exact match first (highest priority)
        if normalized in alias_lookup and alias_lookup[normalized] not in resolved:
            resolved[alias_lookup[normalized]] = idx
            continue
        # Partial/substring match fallback
        for alias_norm, canonical in alias_lookup.items():
            if alias_norm in normalized and canonical not in resolved:
                resolved[canonical] = idx
    return resolved


def _merge_split_header(all_rows: list[tuple], header_row_idx: int) -> tuple:
    """Merge two-row split headers (common in Tally Voucher Register exports)."""
    header_row = all_rows[header_row_idx]
    if header_row_idx + 1 >= len(all_rows):
        return header_row

    next_row = all_rows[header_row_idx + 1]
    continuation_cells = 0
    for cell in next_row:
        if cell is None or cell == "":
            continue
        if not isinstance(cell, str):
            return header_row   # data row, not a header continuation
        if len(cell.strip()) > 25:
            return header_row
        continuation_cells += 1

    if continuation_cells == 0:
        return header_row

    merged = []
    for i, cell in enumerate(header_row):
        next_cell = next_row[i] if i < len(next_row) else None
        if next_cell and isinstance(next_cell, str) and next_cell.strip():
            merged.append((safe_str(cell) + " " + next_cell.strip()).strip())
        else:
            merged.append(cell)
    return tuple(merged)


# ── Shared row-parsing logic ──────────────────────────────────────────────────

def _parse_rows(all_rows: list[tuple], file_name: str) -> Gstr2AParseResult:
    """Parse normalized rows (from openpyxl or xlrd) into structured output."""

    # Extract file-level metadata from first 20 rows
    metadata = Gstr2AMetadata()
    for row in all_rows[:20]:
        if not row:
            continue
        left  = safe_str(row[0]).lower() if len(row) > 0 else ""
        right = safe_str(row[1])         if len(row) > 1 else ""

        if left.startswith("company") and right:
            metadata.company_name = right
        elif left.startswith("period") and right and " to " in right:
            parts = right.split(" to ", 1)
            metadata.period_start = parts[0].strip()
            metadata.period_end   = parts[1].strip()
        elif left.startswith("gst registration") and right:
            metadata.gstin = right
        elif not metadata.company_name and row[1] is None and left and not left.startswith("voucher"):
            metadata.company_name = safe_str(row[0])
        elif not metadata.period_start and " to " in left:
            parts = left.split(" to ", 1)
            metadata.period_start = parts[0].strip()
            metadata.period_end   = parts[1].strip()

    # Detect + build column map
    header_row_idx = _detect_header_row(all_rows)
    if header_row_idx == -1:
        raise ValueError(f"Could not find header row in GSTR-2A file: {file_name}")

    merged_header  = _merge_split_header(all_rows, header_row_idx)
    resolved       = _build_column_map(merged_header)
    has_split      = merged_header != all_rows[header_row_idx]
    data_start_idx = header_row_idx + (2 if has_split else 1)

    # Validate minimum required columns
    missing_core = [c for c in CORE_COLUMNS if c not in resolved]
    if missing_core:
        raise ValueError(
            f"Missing required columns in GSTR-2A file '{file_name}': {missing_core}. "
            "Expected at least: GSTIN, Invoice Date, Taxable Amount."
        )

    has_vch_col     = "vch_number"     in resolved
    has_inv_col     = "invoice_number" in resolved

    invoices: list[Gstr2AInvoice] = []

    for row in all_rows[data_start_idx:]:
        if not row or all(cell is None for cell in row):
            continue

        def get(col: str):
            idx = resolved.get(col)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        particulars = safe_str(get("particulars"))
        if "total" in particulars.lower():
            continue

        date_val = get("invoice_date")
        if date_val is None or safe_str(date_val) == "":
            continue

        party_gstin = safe_str(get("party_gstin")).upper().replace(" ", "")
        if not party_gstin:
            continue

        # ── Resolve both invoice number fields ────────────────────────────────
        vch_no_val         = safe_str(get("vch_number"))     if has_vch_col else ""
        invoice_number_val = safe_str(get("invoice_number")) if has_inv_col else ""

        # Cross-fill when one is missing
        if not invoice_number_val:
            invoice_number_val = vch_no_val    # older export: use vch_no for both
        if not vch_no_val:
            vch_no_val = invoice_number_val    # symmetrical

        if not vch_no_val and not invoice_number_val:
            continue  # no invoice reference at all — skip

        # ── Tax fields ────────────────────────────────────────────────────────
        taxable_amount = safe_number(get("taxable_amount"))
        igst           = safe_number(get("igst"))
        cgst           = safe_number(get("cgst"))
        sgst           = safe_number(get("sgst"))
        cess_val       = safe_number(get("cess"))
        tax_amount     = safe_number(get("tax_amount"))
        if tax_amount == 0.0:
            tax_amount = round(igst + cgst + sgst + cess_val, 2)

        raw_inv_amount = get("invoice_amount")
        invoice_amount = (
            safe_number(raw_inv_amount)
            if raw_inv_amount is not None
            else taxable_amount + tax_amount
        )

        invoices.append(Gstr2AInvoice(
            date           = safe_str(date_val),
            particulars    = particulars,
            party_gstin    = party_gstin,
            vch_type       = safe_str(get("vch_type")),
            vch_no         = vch_no_val,
            invoice_number = invoice_number_val,
            taxable_amount = taxable_amount,
            igst           = igst,
            cgst           = cgst,
            sgst_utgst     = sgst,
            cess           = cess_val,
            tax_amount     = tax_amount,
            invoice_amount = invoice_amount,
        ))

    return Gstr2AParseResult(metadata=metadata, invoices=invoices)


# ── Format-specific readers ───────────────────────────────────────────────────

def _parse_xlsx(file_bytes: bytes, file_name: str) -> Gstr2AParseResult:
    """Read .xlsx using openpyxl (xlrd v2 does NOT support .xlsx)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws       = wb.worksheets[0]
        all_rows = list(ws.iter_rows(values_only=True))
        return _parse_rows(all_rows, file_name)
    finally:
        wb.close()


def _xls_cell_value(ws, row_idx: int, col_idx: int, datemode: int):
    """
    Convert an xlrd cell to a Python value.
    xlrd returns date cells as Excel serial floats — convert to formatted string.
    """
    cell = ws.cell(row_idx, col_idx)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            t  = xlrd.xldate_as_tuple(cell.value, datemode)
            dt = datetime(*t[:6])
            return dt.strftime("%-d-%b-%y")
        except Exception:
            return safe_str(cell.value)
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return None
    return cell.value


def _parse_xls(file_bytes: bytes, file_name: str) -> Gstr2AParseResult:
    """Read legacy .xls using xlrd with proper date-serial conversion."""
    wb       = xlrd.open_workbook(file_contents=file_bytes)
    ws       = wb.sheet_by_index(0)
    datemode = wb.datemode
    all_rows = [
        tuple(_xls_cell_value(ws, i, j, datemode) for j in range(ws.ncols))
        for i in range(ws.nrows)
    ]
    return _parse_rows(all_rows, file_name)


# ── Public entry point ────────────────────────────────────────────────────────

def parse_gstr2a(file_bytes: bytes, file_name: str) -> Gstr2AParseResult:
    """
    Parse a GSTR-2A Excel file and return structured data.

    Routing (IMPORTANT — xlrd v2 does NOT support .xlsx):
      .xlsx or .xlsm  → openpyxl
      .xls            → xlrd
      unknown         → try openpyxl first, fall back to xlrd
    """
    name_lower = file_name.lower()

    if name_lower.endswith(".xlsx") or name_lower.endswith(".xlsm"):
        return _parse_xlsx(file_bytes, file_name)

    if name_lower.endswith(".xls"):
        # Try xlrd for genuine binary .xls files.
        # If it fails, the file is likely xlsx-format with .xls extension (Tally).
        try:
            return _parse_xls(file_bytes, file_name)
        except Exception:
            return _parse_xlsx(file_bytes, file_name)

    # Unknown extension — try openpyxl first, then xlrd
    try:
        return _parse_xlsx(file_bytes, file_name)
    except Exception:
        return _parse_xls(file_bytes, file_name)